import openpyxl
import time
import os
import win32com.client as win32
import pandas as pd
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import xlsxwriter


file_path ='C:\\work\\sunyong\\test'
dbcsv_path ='C:\\work\\sunyong\\test\\tt.csv'
dbxlsx_path = 'C:\\work\\sunyong\\test\\tt2.xlsx'
rpa_xlsx_path = 'C:\\work\\sunyong\\test\\TEST230221-4-TEST_품목사항.xls.xlsx'



#xlx파일을 xlsx로 바꾸기 위한 함수
def xls2xlsx(file):
    #xlsx파일이 있는지 여부 있으면 삭제
    if os.path.isfile(file + "x"):
        os.remove(file + "x")

    wb = win32.gencache.EnsureDispatch('Excel.Application').Workbooks.Open(file)

    #끝에 x를 붙여 xlsx파일로 저장하게 한다.
    wb.SaveAs(file + "x", FileFormat=51)
    wb.Close()

    win32.gencache.EnsureDispatch('Excel.Application').Application.Quit()
    time.sleep(3)

    xlsx_path = file + "x"

    return xlsx_path

def other(a):
    x = sum(a, [])
    sw = list(dict.fromkeys(x))
    if len(sw) == 3:
        return sw
    nw = [s for s in sw if s == s]
    return nw



def main(file):
    #xls파일을 xlsx로 변환
    xls2xlsx(file)
    rpa_xlsx_path = xls2xlsx(file)
    #csv파일을 xlsx로 변환
    r_csv = pd.read_csv(dbcsv_path, encoding='UTF8')
    save_xlsx = pd.ExcelWriter(dbxlsx_path, engine='xlsxwriter')
    r_csv.to_excel(save_xlsx, index=False)
    save_xlsx.save()
    save_xlsx.close()
    #다운로드 받은 RPA excel파일에서 품목코드 가져오기
    wb_or = openpyxl.load_workbook(rpa_xlsx_path)
    ws_or = wb_or.active
    pummok_code = []
    for i in range(4, ws_or.max_row+1):
        pummok= ws_or.cell(row=i, column=5).value
        if pummok is None:
            pass
        else:
            pummok_code.append(pummok)
    #db파일에서 품목코드 가져오기
    wb_tt = openpyxl.load_workbook(dbxlsx_path)
    ws_tt = wb_tt.active
    pummok_code_tt=[]
    for j in range(2, ws_tt.max_row+1):
        pummok_tt = ws_tt.cell(row=j, column =4).value
        if pummok_tt is None:
            pass
        else:
            pummok_code_tt.append(pummok_tt)
    #db파일과 RPA파일에서 중복된 것들만 가져와서 리스트에 담기
    pummok_dup =list(set(pummok_code_tt).intersection(pummok_code))
    #품목코드를 하나씩 가져와 db서 해당하는 품목코드의 매입단가와 매입코드 가져오기
    r_tt = pd.read_excel(dbxlsx_path)
    for data in pummok_dup:
        #data에 해당하는 품목코드의 dataframe 값들만 가져오기
        df = r_tt[r_tt['item_code'] == data]
        #db서 가져온 품목코드중 최소값의 매입단가 가져오기
        purchase_min = df['purchaser_price'].min()
        b = df[df['purchaser_price']==purchase_min]
        # 최저 매입단가, 매입코드, 품목코드를 리스트로 출력
        a = b.values.tolist()

        if len(a) >= 2:
            other(a)
            new = other(a)
            code = new[3]
            for z in range(2, ws_or.max_row+1):
                #RPA excel에 해당코드면 값넣기
                if ws_or.cell(row=z, column=5).value == code:
                    ws_or['Q'+str(z)] = new[2]
                    ws_or['R'+str(z)] = new[1]
                    wb_or.save('C:\\work\\sunyong\\test\\result.xlsx')
        else:
            # 이중 리스트이기 때문에 일반 리스트로 변환하기
            input_data = a[0]
            # 리스트에서 품목코드 가져오기
            code = input_data[3]
            for z in range(2, ws_or.max_row+1):
                #RPA excel에 해당코드면 값넣기
                if ws_or.cell(row=z, column=5).value == code:
                    ws_or['Q'+str(z)] = input_data[2]
                    ws_or['R'+str(z)] = input_data[1]
                    wb_or.save('C:\\work\\sunyong\\test\\result.xlsx')
    print(pummok_dup)
    ws_or.delete_rows(1, 2)
    ws_or.delete_cols(2, 15)
    ws_or.delete_cols(4)
    for z in range(2, ws_or.max_row + 1):
        if ws_or['A' + str(z)].value == '계:':
            ws_or.delete_rows(z)
        wb_or.save('C:\\work\\sunyong\\test\\result.xlsx')

    if not pummok_dup:
        ws_or.delete_rows(1, 2)
        ws_or.delete_cols(2, 15)
        ws_or.delete_cols(4)
        for z in range(2, ws_or.max_row+1):
            if ws_or['A' + str(z)].value == '계:':
                ws_or.delete_rows(z)
            wb_or.save('C:\\work\\sunyong\\test\\result.xlsx')

    wb_re = openpyxl.load_workbook('C:\\work\\sunyong\\test\\result.xlsx')
    ws_re = wb_re.active

    wb_nre = openpyxl.load_workbook('C:\\work\\sunyong\\test\\new_result.xlsx')
    ws_nre = wb_nre.active


    for i in range(2, ws_re.max_row+1):
        ws_nre['A'+str(i)].value = ws_re['A'+str(i)].value
        ws_nre['B'+str(i)].value = ws_re['B'+str(i)].value
        ws_nre['C'+str(i)].value = ws_re['C'+str(i)].value
    wb_nre.save('C:\\work\\sunyong\\test\\new_result.xlsx')



    # #처리완료 파일 삭제하기
    # a = os.listdir(file_path)
    # b = [file for file in a if file.endswith('.xlsx')]
    # for x in b:
    #     print(x)
    #     os.remove('C:\\work\\sunyong\\test' + '\\' + x)

if __name__ == "__main__":
    main(r'C:\work\sunyong\test\TEST230221-4-TEST_품목사항.xls')
